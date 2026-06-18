"""XLSX reporting for daily application runs."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelReporter:
    """Writes daily application and outreach records to an Excel workbook."""

    HEADERS = [
        "Date/Time",
        "Platform",
        "Job Title",
        "Company",
        "Employer Email",
        "Location",
        "Job URL",
        "Apply Type",
        "Match Score",
        "Matched Skills",
        "Generated Subject",
        "Generated Text Preview",
        "Application Status",
        "Outreach Status",
        "Failure/Manual Reason",
    ]

    def __init__(self, reports_dir: Path):
        self.reports_dir = reports_dir

    def write_daily(self, rows: list[dict], date_slug: str) -> Path:
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        path = self.reports_dir / f"applications_{date_slug}.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Applications"
        sheet.append(self.HEADERS)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append(self._row_values(row))

        for idx, _ in enumerate(self.HEADERS, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = 22

        workbook.save(path)
        return path

    def _row_values(self, row: dict) -> list:
        body = row.get("body") or ""
        return [
            row.get("applied_at") or row.get("sent_at") or "",
            row.get("platform") or "",
            row.get("title") or "",
            row.get("company") or "",
            row.get("recipient") or "",
            row.get("location") or "",
            row.get("job_url") or "",
            self._apply_type(row),
            row.get("match_score") or "",
            self._keywords(row.get("matched_keywords")),
            row.get("subject") or "",
            body[:240],
            row.get("application_status") or "",
            row.get("outreach_status") or "",
            row.get("application_reason")
            or row.get("outreach_reason")
            or row.get("safety_reason")
            or "",
        ]

    @staticmethod
    def _apply_type(row: dict) -> str:
        if row.get("is_easy_apply"):
            return "Easy Apply"
        if row.get("is_external"):
            return "External"
        return ""

    @staticmethod
    def _keywords(value) -> str:
        if not value:
            return ""
        if isinstance(value, list):
            return ", ".join(value)
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return ", ".join(str(item) for item in parsed)
        except Exception:
            return str(value)
        return str(value)
